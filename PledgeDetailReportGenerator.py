#   XLSX Report Generator
#   Copyright 2018 by Sean Vo Kirkpatrick using GNU GPL v3
#   skirkpatrick@racc.org or sean@studioantipode.com or seanvokirkpatrick@gmail.com
#   
#   For each Appeal ID group in the input file, open (or create if it doesn't exist) the proper Excel file
#   and append a sheet with a report for the details associated with that Appeal ID group.
#   This program uses the openpyxl library, both for its ability to append sheets to existing Excel files and
#   to show a different method of Excel report generation (the Check Report Generator uses the xlsxwriter library)
#
#   Tested using    - Anaconda 5.0.0
#                   - pandas 0.22.0
#                   - openpyxl 2.5.1
#   IDE: Visual Studio 2017 Community Edition

# License Info:
#   This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

import datetime as dt
from openpyxl.worksheet.header_footer import HeaderFooterItem
from openpyxl.worksheet.header_footer import HeaderFooter
import os
import pandas as pd
import pyodbc
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Protection, Font, Color, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
#from openpyxl.styles.numbers import FORMAT_DATE_XLSX14
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
import numpy as np

def format_range(r1,c1,r2,c2,**options):
    for column in range(c1,c2+1):
        for row in range(r1,r2+1):
            _cell = ws.cell(row=row,column=column)
            if options.get("fill"):
                _cell.fill = options.get("fill")
            if options.get("font"):
                _cell.font = options.get("font")
            if options.get("num_format"):
                _cell.number_format = options.get("num_format")
    return

def write_detailrow(r,row,ws,column_widths):
    
    ws.cell(row=r,column=1,value="{0}".format(row[7]))          #Name
    ws.cell(row=r,column=2,value="{0}".format(row[11]))          #Reference
    ws.cell(row=r,column=3,value="{0}".format(row[3]))          #Appeal ID
    ws.cell(row=r,column=4,value="{0}".format(row[18]))         #Date
    ws.cell(row=r,column=5,value="{0}".format(row[2]))         #Fund (description)
    ws.cell(row=r,column=6,value="{0}".format(row[5]))         #Gift ID (from Raiser's Edge)
    _a = ws.cell(row=r,column=7,value="={0}".format(row[4]))   #Amount
    _a.number_format = FORMAT_CURRENCY_USD_SIMPLE   #set formatting to USD currency standard
    for i in range(1,8):
        if len(ws.cell(row=r,column=i).value) > column_widths[i-1]:
            column_widths[i-1] = len(ws.cell(row=r,column=i).value)
    r+=1
    return(r,column_widths)

def write_summaryrow(r,_r,ws,category):
    ws.cell(row=r,column=1,value="Subtotal - {0}".format(category))
    _a = ws.cell(row=r,column=7,value="=SUBTOTAL(109,G{0}:G{1})".format(_r,r-1))
    _a.number_format = FORMAT_CURRENCY_USD_SIMPLE
    font = Font(b=True)
    fill = PatternFill(fill_type='solid',patternType='solid',fgColor=Color(rgb="00b7b7b7"))    #light grey
    format_range(r,1,r,7,fill=fill,font=font)
    r+=1
    return(r)

def write_totalrow(r,_r,ws):
    ws.cell(row=r,column=1,value="Total")
    _a = ws.cell(row=r,column=7,value="=SUBTOTAL(109,G{0}:G{1})".format(_r,r-2))
    _a.number_format = FORMAT_CURRENCY_USD_SIMPLE
    for c in range(1,8):
        _a = ws.cell(row=r,column=c)
        _a.font = Font(b=True,color=Color(rgb="00FFFFFF"))
        _a.fill = PatternFill(fill_type='solid',patternType='solid',fgColor=Color(rgb="00000000"))
    r+=1
    return(r)


#-------------------------------------------------------------------------------
# Input:  Select the date range (for Gift Date) that you wish to create reports for
startdate = '2018-12-05'         
enddate = '2018-12-31'
#-------------------------------------------------------------------------------

#database connections
#connect to db, requires windows integrated security for this connection string
cnxn = pyodbc.connect("Driver={SQL Server Native Client 11.0};" #requires explicitily stating the sql driver
                      "Server=overlook;"
                      "Database=re_racc;"
                      "Trusted_Connection=yes;")    #use windows integrated security
cursor = cnxn.cursor()

sqlcommand = 'exec sp_pledgereport ''?'', ''?'' '  #call stored procedure
#note that the FundCategory sorting is done in the stored procudure.  the order is:
# ["Designated", "Arts Ed", "Right Brain", "RACC", "Community", "Holding", "Undesignated"]
sqlparams = (startdate,enddate)
cursor.execute(sqlcommand,sqlparams)
columns = [column[0] for column in cursor.description]

data = []   #grab results, put into a list, put list into numpy array, and then put numpy array into pandas dataframe
for row in cursor:
    data.append(tuple(row))
df = pd.DataFrame.from_records(np.array(data),columns=columns)
df = df.replace(pd.np.nan, '', regex=True)  #replaces all types of NAN entries with blank space
print("Query results count(rows): " + str(df.shape[0]))

filepath = "C:\\Users\\skirkpatrick\\Coding\\Python\\"
#outputpath = "\\\\CONCORDIA\\lancentral\\Work for Art\\Raisers Edge Reports\\Pledge reports\\17-18\\"
outputpath = filepath

# columns of stored procedure:
# Campaign,Fund,Appeal ID,Amount,GiftID,Constituent ID,Name,FIRST_NAME,KEY_NAME,LAST_NAME,Reference,PRIMARY_ADDRESSEE,PRIMARY_SALUTATION,ADDRESS_BLOCK,CITY,STATE,POST_CODE,Gift Date,Fund Category,FUND_ID

groupedby_Appeal = df.groupby('Appeal ID')

#format definitions
font_header = Font(b=True,color=Color(rgb="00FFFFFF"))
fill_header = PatternFill(fill_type="solid",fgColor=Color(rgb="00434343"))
font_header2 = Font(b=True,color=Color(rgb="00FFFFFF"))
fill_header2 = PatternFill(fill_type="solid",fgColor=Color(rgb="00434343"))

for name1, group1 in groupedby_Appeal:
    #try opening file, replacing '/' with '-' to satisfy file naming rules
    fp = outputpath +  "_" + str(name1).replace("/","-") + ".xlsx"
    #fp2 = outputpath + "Test\\" + str(name1).replace("/","-") + "_test.xlsx"
    #check if [Appeal ID].xlsx exists
    if os.path.exists(fp):
        print("{0} exists!".format(name1))
        wb = openpyxl.load_workbook(fp)
        ws = wb.create_sheet()
    else:
        print("{0} not found! Creating new file...".format(fp))
        wb = Workbook()
        ws = wb.active

    ws.title = "{0}".format(dt.datetime.now().strftime("%m-%d-%y"))
    column_widths = [19,10,12,11,20,11,11]   #initial guesses at column width
    
    #Page Setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.HeaderFooter.differentOddEven = False    
    ws.HeaderFooter.oddHeader.center.text = "Pledge Report"
    ws.HeaderFooter.oddFooter.left.text = "&Z&F"
    ws.HeaderFooter.oddFooter.right.text = "Page &P of &N"
    
    #Rollup report for top of report
    #print list of categories
    r=2 #openpyxl uses 1-based index, same as excel
    ws.cell(row=1,column=5,value="Category")
    ws.cell(row=1,column=6,value="Subtotals")
    ws.cell(row=1,column=7,value="Totals")
    format_range(1,5,1,7,font=font_header2,fill=fill_header2)
    startingrolluprow = 2
    r2=2    #row counter #2

    groupby_category = group1.groupby('Fund Category')
    for n1,g1 in groupby_category:
        ws.cell(column=5,row=r,value=n1)
        r += 1
    startingdatarow = r + 3     #indicates which row to start writing data to
    if(startingdatarow < 4):
        startingdatarow = 4
    r = startingdatarow
    #writer header
    ws.cell(row=startingdatarow-1,column=1,value="Name")
    ws.cell(row=startingdatarow-1,column=2,value="Reference")
    ws.cell(row=startingdatarow-1,column=3,value="Appeal ID")
    ws.cell(row=startingdatarow-1,column=4,value="Date")
    ws.cell(row=startingdatarow-1,column=5,value="Fund")
    ws.cell(row=startingdatarow-1,column=6,value="Gift ID")
    ws.cell(row=startingdatarow-1,column=7,value="Amount")
    format_range(startingdatarow-1,1,startingdatarow-1,7,font=font_header,fill=fill_header)
    ws.print_title_rows = "{0}:{1}".format(startingdatarow-1,startingdatarow-1)     #repeat row at top of each page

    for n1,g1 in groupby_category:
        startingcategoryrow = r
        cat =""     #placeholder for scope reasons
        for row in g1.itertuples():
            print(row)
            r,column_widths = write_detailrow(r,row,ws,column_widths)
            cat = row[2]    #retain category name from the row to use in the subtotal description
        r = write_summaryrow(r,startingcategoryrow,ws,cat)
        ws.cell(row=r2,column=6,value="=G{0}".format(r-1))
        r2+=1
    r = write_totalrow(r,startingdatarow,ws)
    _a = ws.cell(row=r2,column=7,value="=G{0}".format(r-1))
    format_range(startingrolluprow,6,r2,7,num_format=FORMAT_CURRENCY_USD_SIMPLE)

    #resize columns
    for i in range(1,8):
        ws.column_dimensions[get_column_letter(i)].width = column_widths[i-1]
    ws.cell(row=r+2,column=1,value="Reported by Sean K. on {}".format(dt.datetime.now().strftime("%m/%d/%y")))
    wb.save(fp)